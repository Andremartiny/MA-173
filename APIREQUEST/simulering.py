import requests
import openpyxl

workbook = openpyxl.load_workbook("C:\\Users\\andrem\\OneDrive - Universitetet i Agder\\Undervisning\\MA-173\\Ukeplan med oppgaver og vedlegg\\Datamateriale\\Kandidat1Mal.xlsx")

# Replace with your own access token and Canvas domain
API_KEY = "10742~bKwu47TRaxvDybBOex52yNcghKRWDtT7gNRxjzXIk4iRWH7WhmAUWRuv0Y9M23UT"
CANVAS_DOMAIN = "uia.instructure.com"

headers = {
    "Authorization": f"Bearer {API_KEY}"
}

# Replace with your desired course ID
course_id = 9847




############################## DELETE ALL ASSIGNMENTS ######################################
def get_assignments():
    response = requests.get(
        f"https://{CANVAS_DOMAIN}/api/v1/courses/{course_id}/assignments?per_page=100",
        headers=headers,
    )
    return {assignment["name"]: assignment["id"] for assignment in response.json()}

def get_quizzes():
    response = requests.get(f"https://{CANVAS_DOMAIN}/api/v1/courses/{course_id}/quizzes?per_page=100",
        headers=headers,
    )
    
    return {quizzes["title"]: quizzes["id"] for quizzes in response.json()}

def delete_assignment(assignment_id):
    response = requests.delete(
        f"https://{CANVAS_DOMAIN}/api/v1/courses/{course_id}/quizzes/{assignment_id}",
        headers=headers,
    )
    return response.status_code == 200
# Replace with the list of assignment names you want to delete
# assignments_to_delete = assignment_names

# Fetch all assignments for the course
# assignments = get_assignments()

# print(assignments)
# Iterate through the list of assignment names and delete each one
# for assignment_ID in assignments:
#     assignment_name = assignments[assignment_ID]
#     if assignment_ID:
#         success = delete_assignment(assignment_ID)
#         if success:
#             print(f"Deleted assignment '{assignment_name}'")
#         else:
#             print(f"Failed to delete assignment '{assignment_name}'")
#     else:
#         print(f"Assignment '{assignment_name}' not found")

###########################################################################################


# ################################### LAGE OPPGAVER ##########################################

def create_assignment(name, group_id=None):
    assignment_data = {
        "quiz": {
            "title": name,
            "submission_types": ["online_upload"],
            "allowed_extensions": ["docx", "pdf", "jpeg", "jpg", "png"],
            "points_possible": 3,
            "grading_type": "points",
            "lock_at": None,
            "allowed_attempts": -1,
            "quiz_type": "assignment",
            "scoring_policy": "keep_highest",
            "published": True,
        }
    }

    if group_id:
        assignment_data["quiz"]["assignment_group_id"] = group_id

    response = requests.post(
        f"https://{CANVAS_DOMAIN}/api/v1/courses/{course_id}/quizzes",
        json=assignment_data,
        headers=headers,
    )
    return response.json()

def get_assignment_groups():
    url = f"https://{CANVAS_DOMAIN}/api/v1/courses/{course_id}/assignment_groups"
    headers = {"Authorization": f"Bearer {API_KEY}"}

    response = requests.get(url, headers=headers)

    if response.status_code == 200:
        return response.json()
    else:
        print(f"Error: {response.status_code}")
        return None

# assignment_groups = {}
# for group in get_assignment_groups():
#     assignment_groups[group["name"]] = group["id"]
# for row in workbook["Datamateriale"].iter_rows(min_row=2, values_only=True):
#     if row[0] == None:
#         continue
#     assignment_group_id = assignment_groups[str(row[0])]
#     assignment_name = str(row[6])
#     print(create_assignment(assignment_name, group_id =  assignment_group_id))


## Steg 1: Lag dict
user_data_tema_lm_user_score = {}
user =59544

## Steg 2: Fyll første nivå med tema --> {'Tallteori : {}, 'Funksjoner' : {}...}
for row in workbook["Datamateriale"].iter_rows(min_row=2, values_only=True):
    if row[0] == None or user_data_tema_lm_user_score.get(str(row[0])) != None:
        continue
    user_data_tema_lm_user_score[row[0]] = {}

## Steg 2: Fyll andre nivå med med LM --> {'Tallteori : {LM1: {}, LM2 : {}}, 'Funksjoner' : {LM1: {}, LM2 : {}}...}
for row in workbook["Datamateriale"].iter_rows(min_row=2, values_only=True):
    if user_data_tema_lm_user_score.get(str(row[0])) != None:
        user_data_tema_lm_user_score[str(row[0])][str(row[6])] = {}

assignment_groups = get_assignment_groups()
quizzes = get_assignments()

## Steg 3: Fyll LM med {userid : {'posted_grade' : score}} --> {'Tallteori : {LM1: {}, LM2 : {}}, 'Funksjoner' : {LM1: {}, LM2 : {}}...}, dermed er hvert LM en dict av alle brukerne i det LM og deres score.
### Denne må endres når en kjører gjennom alle
for row in workbook["Datamateriale"].iter_rows(min_row=2, values_only=True):
    if user_data_tema_lm_user_score.get(str(row[0])) != None:
        user_id = user
        user_data_tema_lm_user_score[str(row[0])][str(row[6])][user_id] = {'posted_grade' : row[5]}

print(user_data_tema_lm_user_score)




for tema in user_data_tema_lm_user_score:
    for assignments in assignment_groups:
        if assignments["name"] == tema:
            tema_id = assignments["id"]
        else:
            continue
    for lm in user_data_tema_lm_user_score[tema]:
        assignment_id = quizzes[lm]
        url = f"https://{CANVAS_DOMAIN}/api/v1/courses/{course_id}/assignments/{assignment_id}/submissions/update_grades"
        data = {"grade_data" : user_data_tema_lm_user_score[tema][lm]}
        response = requests.post(url, json=data, headers=headers)

        if response.status_code == 200 or response.status_code == 201:
            print(f"Successfully updated assignment scores for multiple users")
            print(response.json())
        else:
            print(f"Failed to update assignment scores for multiple users, status code: {response.status_code}, response: {response.text}")
    # print(user_data_tema_lm_user_score[tema])

# #####################################################################################################


# #############################OPPDATERE SCORE######################################
# def get_all_pages(url, headers):
#     all_data = []
#     page = 1

#     while True:
#         response = requests.get(f"{url}&page={page}", headers=headers)

#         if response.status_code == 200:
#             data = response.json()

#             if not data:
#                 break

#             all_data.extend(data)
#             page += 1
#         else:
#             print(f"Failed to fetch data from page {page}, status code: {response.status_code}")
#             break

#     return all_data


# def get_assignment_id(assignment_name):
#     # Send a GET request to the Canvas API to retrieve the assignments list
#     url = f"https://{CANVAS_DOMAIN}/api/v1/courses/{course_id}/assignments?per_page=100"
#     assignments_data = get_all_pages(url, headers)

#     # Iterate through the list of assignments and find the one with the matching name
#     for assignment in assignments_data:
#         if assignment["name"] == assignment_name:
#             print(f"Successfully found assignment '{assignment_name}' with ID {assignment['id']}")
#             return assignment["id"]

#     print(f"Assignment '{assignment_name}' not found")
#     return None

# def update_assignment_score(assignment_id, user_id, assignment_score):
#     url = f"https://{CANVAS_DOMAIN}/api/v1/courses/{course_id}/assignments/{assignment_id}/submissions/{user_id}"
#     data = {
#         "submission": {
#             "posted_grade": assignment_score
#         }
#     }
#     response = requests.put(url, json=data, headers=headers)

#     if response.status_code == 200 or response.status_code == 201:
#         print(f"Successfully updated assignment score for User ID {user_id}")
#         return response.json()
#     else:
#         print(f"Failed to update assignment score for User ID {user_id}, status code: {response.status_code}, response: {response.text}")
#         return None


# for score in assignment_names:
#     id =get_assignment_id(score[0])
#     update_assignment_score(id , quiz_student_id, score[2])
#####################################################################################################


# user_id = 59544
# assignment_score = 2


# quiz_student_id = 59544
# # Your list of assignment names
# assignment_names_and_user_scores = {}
# for row in workbook["Datamateriale"].iter_rows(min_row=2):
#     if row[0].value != 'Didaktikk' and row[0].value != None:
#         if row[6].value not in assignment_names_and_user_scores:
#             assignment_names_and_user_scores[row[6].value] = {}
#         assignment_names_and_user_scores[row[6].value][str(quiz_student_id)] = {"posted_grade" : row[5].value}
# print(assignment_names_and_user_scores)
# # Bulk update assignment grades

# for LM in assignment_names_and_user_scores.keys():
#     print(LM)
#     assignment_id = get_assignment_id(LM)
#     print(assignment_id)
#     print(assignment_names_and_user_scores[LM])

#     url = f"https://{CANVAS_DOMAIN}/api/v1/courses/{course_id}/assignments/{assignment_id}/submissions/update_grades"
#     data = {
#         "grade_data":  assignment_names_and_user_scores[LM]
#     }

#     response = requests.post(url, json=data, headers=headers)

#     if response.status_code == 200 or response.status_code == 201:
#         print(f"Successfully updated assignment score for User ID {user_id}")
#         print(response.json())
#     else:
#         print(f"Failed to update assignment score for User ID {user_id}, status code: {response.status_code}, response: {response.text}")
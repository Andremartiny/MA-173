import requests
import openpyxl

workbook = openpyxl.load_workbook("C:\\Users\\andrem\\OneDrive - Universitetet i Agder\\Undervisning\\MA-173\\Ukeplan med oppgaver og vedlegg\\Datamateriale\\DataAlleKandidater.xlsx")
workbook1 = openpyxl.load_workbook("C:\\Users\\andrem\\OneDrive - Universitetet i Agder\\Undervisning\\MA-173\\Ukeplan med oppgaver og vedlegg\\Datamateriale\\Kandidat1Mal.xlsx")

# Replace with your own access token and Canvas domain
API_KEY = "10742~BVyt0MYvbYdR4dzoGLNXZriL7wfmwKmYwzj1jyVFwd19i585uILGA4sVmsIJ9hrX"
CANVAS_DOMAIN = "uia.instructure.com"

headers = {
    "Authorization": f"Bearer {API_KEY}"
}

# Replace with your desired course ID
# course_id = 9847
course_id = 12547 # MA-173-ID: 12547



def get_quizzes():
    response = requests.get(f"https://{CANVAS_DOMAIN}/api/v1/courses/{course_id}/quizzes?per_page=100",
        headers=headers,
    )
    
    return {quizzes["title"]: quizzes["id"] for quizzes in response.json()}


def create_assignment(name, group_id=None):
    assignment_data = {
        "quiz": {
            "title": name,
            "submission_types": ["online_upload"],
            "allowed_extensions": ["docx", "pdf", "jpeg", "jpg", "png"],
            "points_possible": 3,
            "grading_type": "points",
            "lock_at": None,
            "allowed_attempts": -1,
            "quiz_type": "assignment",
            "scoring_policy": "keep_highest",
            "published": True,
        }
    }

    if group_id:
        assignment_data["quiz"]["assignment_group_id"] = group_id

    response = requests.post(
        f"https://{CANVAS_DOMAIN}/api/v1/courses/{course_id}/quizzes",
        json=assignment_data,
        headers=headers,
    )
    return response.json()

def get_max_score_for_user(assignment_id, user_id):
    url = f"https://{CANVAS_DOMAIN}/api/v1/courses/{course_id}/assignments/{assignment_id}/submissions/{user_id}?include[]=submission_history"

    response = requests.get(url, headers=headers)
    submission = response.json()

    max_score = 0
    if 'submission_history' in submission:
        for attempt in submission['submission_history']:
            if attempt['score'] is not None and attempt['score'] > max_score:
                max_score = attempt['score']

    return max_score

def get_assignment_groups():
    url = f"https://{CANVAS_DOMAIN}/api/v1/courses/{course_id}/assignment_groups"
    headers = {"Authorization": f"Bearer {API_KEY}"}

    response = requests.get(url, headers=headers)

    if response.status_code == 200:
        return response.json()
    else:
        print(f"Error: {response.status_code}")
        return None

def get_assignment_submissions(assignment_id):
    url = f"https://{CANVAS_DOMAIN}/api/v1/courses/{course_id}/assignments/{assignment_id}/submissions"
    headers = {"Authorization": f"Bearer {API_KEY}"}
    
    jsonresponse = []

    page = 1
    
    while True:

        # Add the "page" parameter to the URL
        params = {
            "per_page": 100,
            "page": page
        }

        response = requests.get(url, headers=headers, params=params)
        assignments = response.json()
    
        # If there are no more pages, break out of the loop
        if len(assignments) == 0:
            break

        # Add the assignments to the dictionary
        jsonresponse.extend(assignments)

        # Increment the page number for the next request
        page += 1

    if response.status_code == 200:
        return jsonresponse
    else:
        print(f"Error: {response.status_code}")
        return None

def get_assignments():
    url = f"https://{CANVAS_DOMAIN}/api/v1/courses/{course_id}/assignments"
    headers = {"Authorization": f"Bearer {API_KEY}"}
    
    assignments_dict = {}

    page = 1

    # Keep looping until we have retrieved all pages of results
    while True:

        # Add the "page" parameter to the URL
        params = {
            "per_page": 100,
            "page": page
        }

        response = requests.get(url, headers=headers, params=params)
        assignments = response.json()

        # If there are no more pages, break out of the loop
        if len(assignments) == 0:
            break

        # Add the assignments to the dictionary
        for assignment in assignments:
            assignments_dict[assignment["name"]] = assignment["id"]

        # Increment the page number for the next request
        page += 1
    return assignments_dict

def get_student_email_id_dict():
    url = f"https://{CANVAS_DOMAIN}/api/v1/courses/{course_id}/users?enrollment_type[]=student&per_page=100"

    response = requests.get(url, headers=headers)
    enrolled_students = response.json()

    student_email_id_dict = {}
    for student in enrolled_students:
        user_id = student['id']
        user_email = student['login_id'].replace("@uia.no", "@student.uia.no")
        student_email_id_dict[user_id] = user_email

    return student_email_id_dict





assignments_dict = get_assignments()
list_of_læringsmål = []
student_epost_dict = get_student_email_id_dict()


## Steg 0: Lag en liste med læringsmålene fra Excel
for row in workbook1["Datamateriale"].iter_rows(min_row=2, values_only=True):
    if row[0] != "Didaktikk" and row[0] != None:
        list_of_læringsmål.append(row[6]) 

## Steg 1: Lag en dict som inneholder studentepostene som key og en dict som value {studentepost: {}}
studenter_i_canvas_dict_med_lm_og_score = {}


for student in student_epost_dict:
    studenter_i_canvas_dict_med_lm_og_score[student] = {}

## Steg 2: 

for assignment in assignments_dict:
    if assignment not in list_of_læringsmål:
        continue
    submissions = get_assignment_submissions(assignments_dict[assignment])
    print(assignment)
    for sub in submissions:
        if sub["entered_score"] == None:
            continue
        if sub["user_id"] not in studenter_i_canvas_dict_med_lm_og_score:
            continue
        if assignment in studenter_i_canvas_dict_med_lm_og_score[sub["user_id"]]:
            studenter_i_canvas_dict_med_lm_og_score[sub["user_id"]][assignment] = max(studenter_i_canvas_dict_med_lm_og_score[sub["user_id"]][assignment], sub["entered_score"])
        else:
            studenter_i_canvas_dict_med_lm_og_score[sub["user_id"]][assignment] = sub["entered_score"]


for user_id in studenter_i_canvas_dict_med_lm_og_score:
    sheet = workbook[student_epost_dict[user_id]]
    for row in sheet.iter_rows(min_row=2, max_row=68, values_only=True):
        if row[6] not in studenter_i_canvas_dict_med_lm_og_score[user_id]:
            # print(f"Student: {student_epost_dict[user_id]}")
            # print(f"{row}")
            # print(f"{row[6]} er ikke registrert i Canvas")
            continue
        score = studenter_i_canvas_dict_med_lm_og_score[user_id][row[6]]
        if score != row[5]:
            print(f"OBSOBSOBS!!! Score i Canvas {score} og score i Excel {row[5]} for læringsmål{row[6]}")
        # print(f"Score i Canvas {score} og score i Excel {row[5]} for læringsmål{row[6]}")



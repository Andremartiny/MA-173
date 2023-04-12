import requests
import openpyxl

workbook = openpyxl.load_workbook("C:\\Users\\andrem\\OneDrive - Universitetet i Agder\\Undervisning\\MA-173\\Ukeplan med oppgaver og vedlegg\\Datamateriale\\DataAlleKandidater.xlsx")
workbook1 = openpyxl.load_workbook("C:\\Users\\andrem\\OneDrive - Universitetet i Agder\\Undervisning\\MA-173\\Ukeplan med oppgaver og vedlegg\\Datamateriale\\Kandidat1Mal.xlsx")

# Replace with your own access token and Canvas domain
API_KEY = "10742~bKwu47TRaxvDybBOex52yNcghKRWDtT7gNRxjzXIk4iRWH7WhmAUWRuv0Y9M23UT"
CANVAS_DOMAIN = "uia.instructure.com"

headers = {
    "Authorization": f"Bearer {API_KEY}"
}

# Replace with your desired course ID
# course_id = 9847
# course_id = 12547 # MA-173-ID: 12547


def get_assignments():
    response = requests.get(
        f"https://{CANVAS_DOMAIN}/api/v1/courses/{course_id}/assignments?per_page=100",
        headers=headers,
    )
    return {assignment["name"]: assignment["id"] for assignment in response.json()}

def get_quizzes():
    response = requests.get(f"https://{CANVAS_DOMAIN}/api/v1/courses/{course_id}/quizzes?per_page=100",
        headers=headers,
    )
    
    return {quizzes["title"]: quizzes["id"] for quizzes in response.json()}


def create_assignment(name, group_id=None):
    assignment_data = {
        "quiz": {
            "title": name,
            "submission_types": ["online_upload"],
            "allowed_extensions": ["docx", "pdf", "jpeg", "jpg", "png"],
            "points_possible": 3,
            "grading_type": "points",
            "lock_at": None,
            "allowed_attempts": -1,
            "quiz_type": "assignment",
            "scoring_policy": "keep_highest",
            "published": True,
        }
    }

    if group_id:
        assignment_data["quiz"]["assignment_group_id"] = group_id

    response = requests.post(
        f"https://{CANVAS_DOMAIN}/api/v1/courses/{course_id}/quizzes",
        json=assignment_data,
        headers=headers,
    )
    return response.json()

def get_assignment_groups():
    url = f"https://{CANVAS_DOMAIN}/api/v1/courses/{course_id}/assignment_groups"
    headers = {"Authorization": f"Bearer {API_KEY}"}

    response = requests.get(url, headers=headers)

    if response.status_code == 200:
        return response.json()
    else:
        print(f"Error: {response.status_code}")
        return None

def get_student_email_id_dict(course_id):
    url = f"https://{CANVAS_DOMAIN}/api/v1/courses/{course_id}/users?enrollment_type[]=student&per_page=100"

    response = requests.get(url, headers=headers)
    enrolled_students = response.json()

    student_email_id_dict = {}
    for student in enrolled_students:
        user_id = student['id']
        user_email = student['login_id']
        student_email_id_dict[user_email] = user_id

    return student_email_id_dict

## Steg 1: Lag dict
user_data_tema_lm_user_score = {}
# user =59544 quiz-student.

## Steg 2: Fyll første nivå med tema --> {'Tallteori : {}, 'Funksjoner' : {}...}
for row in workbook1["Datamateriale"].iter_rows(min_row=2, values_only=True):
    if row[0] == None or user_data_tema_lm_user_score.get(str(row[0])) != None:
        continue
    user_data_tema_lm_user_score[row[0]] = {}

## Steg 2: Fyll andre nivå med med LM --> {'Tallteori : {LM1: {}, LM2 : {}}, 'Funksjoner' : {LM1: {}, LM2 : {}}...}
for row in workbook1["Datamateriale"].iter_rows(min_row=2, values_only=True):
    if user_data_tema_lm_user_score.get(str(row[0])) != None:
        user_data_tema_lm_user_score[str(row[0])][str(row[6])] = {}

assignment_groups = get_assignment_groups()
quizzes = get_assignments()
student_ids = get_student_email_id_dict(course_id)

## Steg 3: Fyll LM med {userid : {'posted_grade' : score}} --> {'Tallteori : {LM1: {}, LM2 : {}}, 'Funksjoner' : {LM1: {}, LM2 : {}}...}, dermed er hvert LM en dict av alle brukerne i det LM og deres score.
### Denne må endres når en kjører gjennom alle
for sheet in workbook.worksheets:
    if sheet.title == "Kandidat1":
        # Process the sheet as needed
        print(f"Processing sheet: {sheet.title}")
        continue
    email = sheet.title
    if email in student_ids:
        email = sheet.title
    elif email.replace("@student.uia.no", "@uia.no") in student_ids:
        email = sheet.title.replace("@student.uia.no", "@uia.no")
    else:
        print(email.replace("@uia.no", "@student.uia.no") + " og " + email + " er ikke blant brukerne i Canvas")
        continue
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if user_data_tema_lm_user_score.get(str(row[0])) != None:
            user_id = student_ids[email]
            user_data_tema_lm_user_score[str(row[0])][str(row[6])][user_id] = {'posted_grade' : row[5]}




for tema in user_data_tema_lm_user_score:
    for assignments in assignment_groups:
        if assignments["name"] == tema:
            tema_id = assignments["id"]
        else:
            continue
    for lm in user_data_tema_lm_user_score[tema]:
        if lm not in quizzes:
            print(f"{lm} er ikke i en av Quizzene i Canvas")
            continue
        assignment_id = quizzes[lm]
        url = f"https://{CANVAS_DOMAIN}/api/v1/courses/{course_id}/assignments/{assignment_id}/submissions/update_grades"
        data = {"grade_data" : user_data_tema_lm_user_score[tema][lm]}
        response = requests.post(url, json=data, headers=headers)

        if response.status_code == 200 or response.status_code == 201:
            print(f"Successfully updated assignment scores for multiple users")
            print(response.json())
        else:
            print(f"Failed to update assignment scores for multiple users, status code: {response.status_code}, response: {response.text}")
 